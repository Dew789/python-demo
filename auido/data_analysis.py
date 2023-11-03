import os
import numpy as np

from openpyxl import Workbook
from app.mechanism_analysis import Analysis
from openpyxl.chart import BarChart, Reference


class DataAnalysis:
    """
    https://zhuanlan.zhihu.com/p/409954748
    """

    def __init__(self, file_name):
        self.file_path = r"D:\mechanism\%s.xlsx" % file_name
        self.wb = Workbook()

    def save(self):
        self.wb.save(self.file_path)

    def find_wav_files(self, path):
        if os.path.isfile(path):
            if path.endswith("wav"):
                yield path
            else:
                print("[%s]非音频文件，忽略解析" % path)

        if os.path.isdir(path):
            for root, ds, fs in os.walk(path):
                for d in ds:
                    yield from self.find_wav_files(os.path.join(root, d))
                for f in fs:
                    yield from self.find_wav_files(os.path.join(root, f))

    def basic_frequency(self, audio_dir, title):
        result = []

        for audio_path in self.find_wav_files(audio_dir):
            freq = Analysis(audio_path).basic_frequency()
            result.append(freq)

        self.draw_count(result, title, "主频（HZ）", "数量")

    def fundamental_frequency_rate(self, audio_dir, title):
        result = []

        for audio_path in self.find_wav_files(audio_dir):
            rate = Analysis(audio_path).fundamental_frequency_rate()
            result.append(round(rate, 3))

        self.draw_histogram(result, title, "基频比重", "数量", 0.1)

    def fifty_multiplier_rate(self, audio_dir, title):
        result = []

        for audio_path in self.find_wav_files(audio_dir):
            rate = Analysis(audio_path).fifty_multiplier_rate()
            rate = round(rate, 2)
            result[rate] = result.get(rate, 0) + 1

        self.draw_histogram(result, title, "50HZ奇偶倍频比重", "数量", 0.2)

    def vibrational_entropy(self, audio_dir, title):
        result = []

        for audio_path in self.find_wav_files(audio_dir):
            rate = Analysis(audio_path).vibrational_entropy(audio_path)
            result.append(round(rate, 2))

        self.draw_histogram(result, title, "振动熵", "数量", 0.5)

    def high_low_rate(self, audio_dir, title):
        result = []

        for audio_path in self.find_wav_files(audio_dir):
            rate = Analysis(audio_path).high_low_rate(audio_path)
            result.append(round(rate, 2))

        self.draw_histogram(result, title, "高低频比", "数量", 0.2)

    def draw_count(self, result, title, x_name, y_name):
        count = {}
        for item in result:
            count[item] = count.get(item, 0) + 1

        ws = self.wb.create_sheet(title)
        ws.append(("主频（HZ）", "数量"))
        for freq, total in sorted(count.items()):
            ws.append((freq, total))

        max_row = len(count) + 1

        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = title
        chart1.x_axis.title = x_name
        chart1.y_axis.title = y_name

        data = Reference(ws, min_col=2, max_col=2,  min_row=1, max_row=max_row)  # 真实数据
        series = Reference(ws, min_col=1, min_row=2, max_row=max_row)            # 横坐标
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(series)
        ws.add_chart(chart1, "D1")

    def draw_histogram(self, result, title, x_name, y_name, gap):
        ws = self.wb.create_sheet(title)

        max_value = int(max(result)) + 1
        hist, bins = np.histogram(result,
                                  np.arange(0, max_value, gap))
        ws.append((x_name, y_name))
        for i in range(len(hist)):
            ws.append(("%s-%s" % (bins[i], bins[i]+gap), hist[i]))

        max_row = len(hist) + 1

        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = title
        chart1.x_axis.title = x_name
        chart1.y_axis.title = y_name

        data = Reference(ws, min_col=2, max_col=2,  min_row=1, max_row=max_row)  # 真实数据
        series = Reference(ws, min_col=1, min_row=2, max_row=max_row)            # 横坐标
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(series)
        ws.add_chart(chart1, "D1")


if __name__ == "__main__":
    data = [
        (r"D:\mechanism\audio\纯变压器声音", "正常运行"),
        (r"D:\mechanism\audio\设备异常样本\1直流偏磁-200条", "直流偏磁"),
        (r"D:\mechanism\audio\设备异常样本\2短路冲击-100条", "短路冲击"),
        (r"D:\mechanism\audio\设备异常样本\3局部放电-450条", "局部放电"),
        (r"D:\mechanism\audio\设备异常样本\4重过载-400条", "重过载"),
        (r"D:\mechanism\audio\设备异常样本\5开关异常-400条", "开关异常"),
        (r"D:\mechanism\audio\环境异音\1鸟鸣-300条", "鸟鸣（正常运行）"),
        (r"D:\mechanism\audio\环境异音\2雨声-300条", "雨声（正常运行）")
    ]

    analysis = DataAnalysis("主频")
    for audio_dir, name in data:
        analysis.basic_frequency(audio_dir, name + "主频")
    analysis.save()

    analysis = DataAnalysis("振动熵")
    for audio_dir, name in data:
        analysis.vibrational_entropy(audio_dir, name + "振动熵")
    analysis.save()

    analysis = DataAnalysis("基频比重")
    for audio_dir, name in data:
        analysis.fundamental_frequency_rate(audio_dir, name + "基频比重")
    analysis.save()

    analysis = DataAnalysis("50HZ奇偶倍频比重")
    for audio_dir, name in data:
        analysis.fifty_multiplier_rate(audio_dir, name + "50HZ奇偶倍频比重")
    analysis.save()

    analysis = DataAnalysis("高低频比")
    for audio_dir, name in data:
        analysis.high_low_rate(audio_dir, name + "高低频比")
    analysis.save()
